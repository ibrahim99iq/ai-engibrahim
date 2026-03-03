import json
import os
from datetime import datetime, date
from openpyxl import load_workbook

EXCEL_PATH = r"C:\Users\ENG-IBRAHIM\OneDrive\Desktop\الداودي\شركة الداودي كاملة.xlsx"
PROJECT_PATH = r"C:\Users\ENG-IBRAHIM\OneDrive\Desktop\ai-engibrahim"

OUT_JSON_PATH = os.path.join(PROJECT_PATH, "data", "api.json")

def fmt_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v)

# اقرأ ناتج المعادلات (لازم الملف محفوظ من Excel)
wb = load_workbook(EXCEL_PATH, data_only=True)
ws = wb["API"]

date_val = ws["A2"].value
total_val = ws["B2"].value
cars_val = ws["C2"].value  # خليه إذا عندك، وإذا فاضي يطلع null

data = {
    "date": fmt_date(date_val),
    "total": total_val,
    "cars": cars_val
}

os.makedirs(os.path.dirname(OUT_JSON_PATH), exist_ok=True)

with open(OUT_JSON_PATH, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("✅ data/api.json updated:", data)